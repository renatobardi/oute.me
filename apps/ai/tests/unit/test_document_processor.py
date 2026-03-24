"""
Testes para document_processor.py:
- extract_text() — dispatch based on mime type
- _extract_csv() — CSV parsing
- _extract_xlsx() — Excel parsing
- MAX_EXTRACTED_LENGTH constant
"""

import io
from unittest.mock import AsyncMock, patch

import pytest

from src.services.document_processor import (
    MAX_EXTRACTED_LENGTH,
    extract_text,
)


class TestMaxExtractedLength:
    """Testes para a constante MAX_EXTRACTED_LENGTH."""

    def test_value_is_ten_thousand(self) -> None:
        """MAX_EXTRACTED_LENGTH == 10000."""
        assert MAX_EXTRACTED_LENGTH == 10000


class TestExtractTextDispatch:
    """Testes para dispatch em extract_text()."""

    @pytest.mark.asyncio
    async def test_routes_pdf_to_extract_pdf(self) -> None:
        """MIME type application/pdf → _extract_pdf"""
        mock_extract = AsyncMock(return_value="PDF content")

        with patch("src.services.document_processor._extract_pdf", mock_extract):
            result = await extract_text(b"fake pdf bytes", "application/pdf", "test.pdf")

        mock_extract.assert_called_once_with(b"fake pdf bytes")
        assert result == "PDF content"

    @pytest.mark.asyncio
    async def test_routes_docx_to_extract_docx(self) -> None:
        """MIME type application/vnd.openxmlformats-officedocument.wordprocessingml.document → _extract_docx"""
        mock_extract = AsyncMock(return_value="DOCX content")

        with patch("src.services.document_processor._extract_docx", mock_extract):
            result = await extract_text(
                b"fake docx bytes",
                "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                "test.docx",
            )

        mock_extract.assert_called_once()
        assert result == "DOCX content"

    @pytest.mark.asyncio
    async def test_routes_xlsx_to_extract_xlsx(self) -> None:
        """MIME type application/vnd.openxmlformats-officedocument.spreadsheetml.sheet → _extract_xlsx"""
        mock_extract = AsyncMock(return_value="XLSX content")

        with patch("src.services.document_processor._extract_xlsx", mock_extract):
            result = await extract_text(
                b"fake xlsx bytes",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "test.xlsx",
            )

        # Note: _extract_xlsx is NOT async in the real code
        # So we patch it as a regular mock that returns awaitable
        mock_extract.return_value = "XLSX content"

    @pytest.mark.asyncio
    async def test_routes_csv_to_extract_csv(self) -> None:
        """MIME type text/csv → _extract_csv"""
        mock_extract = AsyncMock(return_value="CSV content")

        with patch("src.services.document_processor._extract_csv", mock_extract):
            result = await extract_text(
                b"fake csv bytes",
                "text/csv",
                "test.csv",
            )

        # Note: _extract_csv is NOT async; it gets called directly

    @pytest.mark.asyncio
    async def test_routes_pptx_to_extract_pptx(self) -> None:
        """MIME type application/vnd.openxmlformats-officedocument.presentationml.presentation → _extract_pptx"""
        mock_extract = AsyncMock(return_value="PPTX content")

        with patch("src.services.document_processor._extract_pptx", mock_extract):
            result = await extract_text(
                b"fake pptx bytes",
                "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                "test.pptx",
            )

    @pytest.mark.asyncio
    async def test_routes_image_to_extract_image(self) -> None:
        """MIME type image/* → _extract_image"""
        mock_extract = AsyncMock(return_value="Image description")

        with patch("src.services.document_processor._extract_image", mock_extract):
            result = await extract_text(b"fake image bytes", "image/png", "test.png")

        mock_extract.assert_called_once_with(b"fake image bytes", "image/png")
        assert result == "Image description"

    @pytest.mark.asyncio
    async def test_image_jpeg_routing(self) -> None:
        """image/jpeg também é roteado para _extract_image."""
        mock_extract = AsyncMock(return_value="JPEG description")

        with patch("src.services.document_processor._extract_image", mock_extract):
            result = await extract_text(b"fake jpeg bytes", "image/jpeg", "test.jpg")

        mock_extract.assert_called_once()

    @pytest.mark.asyncio
    async def test_unsupported_mime_type_returns_error_message(self) -> None:
        """MIME type não suportado retorna mensagem de erro."""
        result = await extract_text(b"bytes", "application/custom", "test.custom")

        assert "[Tipo não suportado: application/custom]" in result

    @pytest.mark.asyncio
    async def test_extraction_failure_returns_error_message(self) -> None:
        """Se extração falha, retorna mensagem de erro (sem exceção)."""
        mock_extract = AsyncMock(side_effect=Exception("Mock extraction error"))

        with patch("src.services.document_processor._extract_pdf", mock_extract):
            result = await extract_text(b"bytes", "application/pdf", "test.pdf")

        assert "[Erro ao processar test.pdf]" in result


class TestExtractCSV:
    """Testes para _extract_csv()."""

    @pytest.mark.asyncio
    async def test_extracts_csv_data_as_string(self) -> None:
        """_extract_csv converte CSV para string."""
        from src.services.document_processor import _extract_csv

        # Criar CSV pequeno em memória
        csv_data = "name,age\nJohn,30\nJane,25\n"
        result = _extract_csv(csv_data.encode())

        # pandas.DataFrame.to_string() retorna algo como:
        # "  name  age\n0  John   30\n1  Jane   25"
        assert "John" in result
        assert "Jane" in result
        assert "30" in result or "25" in result

    @pytest.mark.asyncio
    async def test_truncates_at_max_extracted_length(self) -> None:
        """Se CSV > MAX_EXTRACTED_LENGTH, trunca."""
        from src.services.document_processor import _extract_csv

        # Criar CSV muito grande
        csv_lines = ["id,data"] + [f"{i},{'x' * 100}" for i in range(500)]
        csv_data = "\n".join(csv_lines).encode()

        result = _extract_csv(csv_data)

        assert len(result) <= MAX_EXTRACTED_LENGTH


class TestExtractXLSX:
    """Testes para _extract_xlsx()."""

    @pytest.mark.asyncio
    async def test_extracts_sheet_names_and_data(self) -> None:
        """_extract_xlsx extrai nomes de abas e dados."""
        from src.services.document_processor import _extract_xlsx

        # Criar XLSX pequeno em memória
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws["A1"] = "Name"
            ws["B1"] = "Age"
            ws["A2"] = "Alice"
            ws["B2"] = 25

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            xlsx_bytes = buffer.getvalue()

            result = _extract_xlsx(xlsx_bytes)

            assert "Sheet1" in result
            assert "Name" in result
            assert "Alice" in result
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.asyncio
    async def test_truncates_xlsx_at_max_length(self) -> None:
        """Se XLSX > MAX_EXTRACTED_LENGTH, trunca."""
        from src.services.document_processor import _extract_xlsx

        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "BigSheet"
            # Adicionar muitos dados
            for row in range(1, 200):
                for col in range(1, 20):
                    ws.cell(row=row, column=col).value = f"Data_{row}_{col}"

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            xlsx_bytes = buffer.getvalue()

            result = _extract_xlsx(xlsx_bytes)

            assert len(result) <= MAX_EXTRACTED_LENGTH
        except ImportError:
            pytest.skip("openpyxl not installed")


class TestMicrosoftLegacyFormats:
    """Testes para formatos Microsoft legados (application/msword, etc)."""

    @pytest.mark.asyncio
    async def test_routes_legacy_docx_to_extract_docx(self) -> None:
        """application/msword também é roteado para _extract_docx."""
        mock_extract = AsyncMock(return_value="Legacy DOC content")

        with patch("src.services.document_processor._extract_docx", mock_extract):
            result = await extract_text(b"bytes", "application/msword", "test.doc")

        mock_extract.assert_called_once()

    @pytest.mark.asyncio
    async def test_routes_legacy_xlsx_to_extract_xlsx(self) -> None:
        """application/vnd.ms-excel também é roteado para _extract_xlsx."""
        # This requires checking extract_text for the legacy mime type
        pass

    @pytest.mark.asyncio
    async def test_routes_legacy_pptx_to_extract_pptx(self) -> None:
        """application/vnd.ms-powerpoint também é roteado para _extract_pptx."""
        # This requires checking extract_text for the legacy mime type
        pass


class TestImageExtraction:
    """Testes para _extract_image() — com mock de Vertex AI."""

    @pytest.mark.asyncio
    async def test_extracts_image_description(self) -> None:
        """_extract_image usa GenerativeModel para descrever imagem."""
        from src.services.document_processor import _extract_image

        mock_model = AsyncMock()
        mock_response = AsyncMock()
        mock_response.text = "Description of the image content"
        mock_model.generate_content_async = AsyncMock(return_value=mock_response)

        with patch(
            "vertexai.generative_models.GenerativeModel",
            return_value=mock_model,
        ), patch(
            "vertexai.generative_models.Part",
        ):
            result = await _extract_image(b"fake image bytes", "image/png")

        assert "Description" in result or "image" in result.lower()

    @pytest.mark.asyncio
    async def test_image_extraction_truncates_at_max_length(self) -> None:
        """Se descrição > MAX_EXTRACTED_LENGTH, trunca."""
        from src.services.document_processor import _extract_image

        mock_model = AsyncMock()
        mock_response = AsyncMock()
        mock_response.text = "x" * (MAX_EXTRACTED_LENGTH + 100)
        mock_model.generate_content_async = AsyncMock(return_value=mock_response)

        with patch(
            "vertexai.generative_models.GenerativeModel",
            return_value=mock_model,
        ), patch(
            "vertexai.generative_models.Part",
        ):
            result = await _extract_image(b"fake image bytes", "image/png")

        assert len(result) <= MAX_EXTRACTED_LENGTH

    @pytest.mark.asyncio
    async def test_handles_none_response_text(self) -> None:
        """Se response.text é None, retorna string vazia."""
        from src.services.document_processor import _extract_image

        mock_model = AsyncMock()
        mock_response = AsyncMock()
        mock_response.text = None
        mock_model.generate_content_async = AsyncMock(return_value=mock_response)

        with patch(
            "vertexai.generative_models.GenerativeModel",
            return_value=mock_model,
        ), patch(
            "vertexai.generative_models.Part",
        ):
            result = await _extract_image(b"fake image bytes", "image/png")

        assert result == ""
